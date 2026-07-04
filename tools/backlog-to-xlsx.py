#!/usr/bin/env python3
"""
BACKLOG.md -> xlsx exporter (a generated VIEW, never committed; BACKLOG.md is the master).

Two ways to run it:
  1. Terminal, from the repo root:   python tools/backlog-to-xlsx.py
  2. Ask Claude Code:                "run the backlog xlsx exporter"

It reads BACKLOG.md at the repo root, parses the pipe table
(ID | Status | Category | Item | Raised | Closed-by), and writes a styled,
filterable spreadsheet to:

    ~/Downloads/threeflows-backlog-<YYYY-MM-DD>.xlsx

The date auto-fills from the system date on each run, so the file self-dates.
Output goes to ~/Downloads -- it is NOT written into the repo and must not be
committed. BACKLOG.md remains the single source of truth.

Requires openpyxl:  pip install openpyxl --break-system-packages
"""
import os
import re
import sys
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("openpyxl is required. Install with: pip install openpyxl --break-system-packages")

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO_ROOT, "BACKLOG.md")
OUT = os.path.join(
    os.path.expanduser("~/Downloads"),
    f"threeflows-backlog-{date.today().isoformat()}.xlsx",
)

HEADERS = ["ID", "Status", "Category", "Item", "Raised", "Closed-by"]
# Status -> cell fill (None = no fill / white). discard also gets strikethrough text.
STATUS_FILL = {
    "close": "C6EFCE",    # green
    "open": None,          # white / none
    "review": "FFEB9C",   # amber
    "park": "D9D9D9",     # grey
    "discard": "FFC7CE",  # light red
}
FONT = "Arial"
COL_WIDTHS = {"A": 9, "B": 10, "C": 14, "D": 95, "E": 12, "F": 30}  # Item (D) widest, wrapped
ROW_RE = re.compile(r"^\|\s*BL-\d+\s*\|")


def parse(path):
    """Parse BL- rows from BACKLOG.md, preserving file order."""
    rows = []
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            if ROW_RE.match(line):
                parts = [c.strip() for c in line.rstrip("\n").split("|")][1:-1]
                if len(parts) != 6:
                    sys.exit(f"Malformed row (expected 6 fields, got {len(parts)}): {parts[:2]}")
                rows.append(parts)
    if not rows:
        sys.exit(f"No BL- rows found in {path}")
    return rows


def build(rows, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    for row in rows:
        ws.append(row)
        r = ws.max_row
        struck = row[1] == "discard"
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10, strike=struck)
            cell.border = border
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=(col == 4))
        fill = STATUS_FILL.get(row[1])
        if fill:
            ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=fill)

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)


def main():
    rows = parse(SRC)
    build(rows, OUT)
    tally = {}
    for row in rows:
        tally[row[1]] = tally.get(row[1], 0) + 1
    print(f"Wrote {OUT}")
    print(f"Rows: {len(rows)}")
    print("Status tally: " + ", ".join(f"{k} {v}" for k, v in sorted(tally.items())))


if __name__ == "__main__":
    main()
